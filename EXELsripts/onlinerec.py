import xlsxwriter
from openpyxl import load_workbook

def write(word, data):
    workbook = xlsxwriter.Workbook('example.xlsx') # name file
    worksheet = workbook.add_worksheet()
    Anum = 1
    A = 'A'
    if len(data) == 0:
        pass
    else:
        for i in range(len(data)):
            a2 = f"{A}{Anum}"
            worksheet.write(a2, data[i])
            Anum += 1
    a2 = f"{A}{Anum}"
    worksheet.write(a2, word)
    workbook.close()

def append(word):
    wb = load_workbook('example.xlsx') #name file
    sheet = wb.get_sheet_by_name('Sheet1') # name list
    anotherSheet = wb.active # IDK
    num = 1
    A = 'A'
    A_data = []
    Searsh = True

    while Searsh:               # достаём
        a = f"{A}{num}"         # все
        a1 = sheet[a].value     # данные
        num +=1                 # из
        if a1 == None:          # колонки
            Searsh = False      # с названием
        else:                   # "А"
            A_data.append(a1)   # и складируем в список

    write(word,A_data)


while True:
    a = input("Что записать?")
    if a == '0':
        break
    else:
        append(a)